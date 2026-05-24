#!/usr/bin/env python3
"""
populate_xlsx.py — write Keyword Research, Page Map, and SEO Briefs tabs from a JSON payload.

Usage:
    python populate_xlsx.py --xlsx SEO_Master.xlsx --payload payload.json

Payload JSON shape:
{
  "keywords": [
    {"keyword": "...", "volume_or_score": 5500, "intent": "Commercial", "funnel": "MOFU",
     "target_page": "Automotive landing", "priority": "High",
     "cpc": 4.62, "serp_features": "AI Overview, People Also Ask, Video carousel",
     "notes": "comp:4, paa"}, ...
  ],
  "pages": [
    {"page": "Automotive landing", "slug": "/for-automotive", "type": "Landing",
     "primary": "...", "secondary_1": "...", "secondary_2": "...", "secondary_3": "...",
     "status": "Draft"}, ...
  ],
  "briefs": [
    {"page": "Automotive landing", "slug": "/for-automotive",
     "title": "...", "meta": "...", "h1": "..."}, ...
  ]
}

Schema contract (see references/xlsx-schema.md):
  - Writes to tabs: Keyword Research, Page Map, SEO Briefs.
  - Never touches tabs: Instructions, Alt Text, Dashboard.
  - Headers live on row 4; data rows start at row 5.
  - SEO Briefs columns D, F, H are formula columns — NEVER written.
  - data_only=False so formulas survive the round trip.

Dependencies: openpyxl
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from typing import Optional

try:
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
except ImportError as e:
    print(f"Missing dependency: {e.name}. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


WRITABLE_TABS = {"Keyword Research", "Page Map", "SEO Briefs"}
PROTECTED_TABS = {"Instructions", "Alt Text", "Dashboard"}

HEADER_ROW = 4
DATA_START_ROW = 5
KW_MAX_ROWS = 100
PAGE_MAX_ROWS = 50
BRIEF_MAX_ROWS = 50

KW_HEADERS = ["Keyword", "Avg. monthly searches", "Intent", "Funnel stage", "Target page", "Priority", "CPC (USD)", "SERP Features", "Notes"]
PAGE_HEADERS = ["Page name", "URL slug", "Page type", "Primary keyword", "Secondary keyword 1", "Secondary keyword 2", "Secondary keyword 3", "Status"]
BRIEF_HEADERS = ["Page", "URL slug", "Title tag", "Title length", "Meta description", "Meta length", "H1", "H1 length"]

BRIEF_FORMULA_COLS = {4, 6, 8}  # 1-indexed columns D, F, H


def _check_headers(ws, expected: list[str], tab_name: str) -> None:
    actual = [ws.cell(row=HEADER_ROW, column=i + 1).value for i in range(len(expected))]
    actual = [(v.strip() if isinstance(v, str) else v) for v in actual]
    if actual != expected:
        raise ValueError(
            f"Tab '{tab_name}' header row {HEADER_ROW} mismatch.\n"
            f"  expected: {expected}\n"
            f"  actual:   {actual}"
        )


def _validate_workbook(wb: Workbook) -> None:
    sheets = set(wb.sheetnames)
    missing_writable = WRITABLE_TABS - sheets
    if missing_writable:
        raise ValueError(f"Workbook is missing writable tab(s): {sorted(missing_writable)}")
    missing_protected = PROTECTED_TABS - sheets
    if missing_protected:
        print(f"Warning: workbook is missing protected tab(s): {sorted(missing_protected)} — continuing.", file=sys.stderr)

    _check_headers(wb["Keyword Research"], KW_HEADERS, "Keyword Research")
    _check_headers(wb["Page Map"], PAGE_HEADERS, "Page Map")
    _check_headers(wb["SEO Briefs"], BRIEF_HEADERS, "SEO Briefs")


def _check_cannibalization(pages: list[dict]) -> None:
    primaries: dict[str, list[str]] = {}
    for p in pages:
        primary = (p.get("primary") or "").strip().lower()
        if not primary:
            continue
        primaries.setdefault(primary, []).append(p.get("page", "<unnamed>"))
    duplicates = {k: v for k, v in primaries.items() if len(v) > 1}
    if duplicates:
        lines = ["Cannibalization detected — primary keywords assigned to multiple pages:"]
        for kw, pages_list in duplicates.items():
            lines.append(f"  '{kw}' → {pages_list}")
        raise ValueError("\n".join(lines))


def _check_cross_refs(keywords: list[dict], pages: list[dict], briefs: list[dict]) -> None:
    kw_set = {(k.get("keyword") or "").strip().lower() for k in keywords}
    page_names = {(p.get("page") or "").strip() for p in pages}

    for p in pages:
        primary = (p.get("primary") or "").strip().lower()
        if primary and primary not in kw_set:
            print(f"Warning: page '{p.get('page')}' primary keyword '{primary}' not in Keyword Research list.", file=sys.stderr)

    for b in briefs:
        if (b.get("page") or "").strip() not in page_names:
            raise ValueError(f"Brief references unknown page: '{b.get('page')}'. Add the page to Page Map first.")


def _clear_data_rows(ws, max_rows: int, col_count: int) -> None:
    for row in range(DATA_START_ROW, DATA_START_ROW + max_rows):
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).value = None


def _clear_briefs_data_rows(ws, max_rows: int) -> None:
    for row in range(DATA_START_ROW, DATA_START_ROW + max_rows):
        for col in range(1, len(BRIEF_HEADERS) + 1):
            if col in BRIEF_FORMULA_COLS:
                continue  # never touch formula columns
            ws.cell(row=row, column=col).value = None


def _write_keywords(ws, keywords: list[dict]) -> int:
    _clear_data_rows(ws, KW_MAX_ROWS, len(KW_HEADERS))
    rows_written = 0
    for i, kw in enumerate(keywords[:KW_MAX_ROWS]):
        row = DATA_START_ROW + i
        ws.cell(row=row, column=1, value=kw.get("keyword", ""))
        ws.cell(row=row, column=2, value=kw.get("volume_or_score"))
        ws.cell(row=row, column=3, value=kw.get("intent", ""))
        ws.cell(row=row, column=4, value=kw.get("funnel", ""))
        ws.cell(row=row, column=5, value=kw.get("target_page", ""))
        ws.cell(row=row, column=6, value=kw.get("priority", ""))
        ws.cell(row=row, column=7, value=kw.get("cpc"))
        ws.cell(row=row, column=8, value=kw.get("serp_features", ""))
        ws.cell(row=row, column=9, value=kw.get("notes", ""))
        rows_written += 1
    return rows_written


def _write_pages(ws, pages: list[dict]) -> int:
    _clear_data_rows(ws, PAGE_MAX_ROWS, len(PAGE_HEADERS))
    rows_written = 0
    for i, p in enumerate(pages[:PAGE_MAX_ROWS]):
        row = DATA_START_ROW + i
        ws.cell(row=row, column=1, value=p.get("page", ""))
        ws.cell(row=row, column=2, value=p.get("slug", ""))
        ws.cell(row=row, column=3, value=p.get("type", ""))
        ws.cell(row=row, column=4, value=p.get("primary", ""))
        ws.cell(row=row, column=5, value=p.get("secondary_1", ""))
        ws.cell(row=row, column=6, value=p.get("secondary_2", ""))
        ws.cell(row=row, column=7, value=p.get("secondary_3", ""))
        ws.cell(row=row, column=8, value=p.get("status", "Draft"))
        rows_written += 1
    return rows_written


def _write_briefs(ws, briefs: list[dict]) -> int:
    _clear_briefs_data_rows(ws, BRIEF_MAX_ROWS)
    rows_written = 0
    for i, b in enumerate(briefs[:BRIEF_MAX_ROWS]):
        row = DATA_START_ROW + i
        ws.cell(row=row, column=1, value=b.get("page", ""))
        ws.cell(row=row, column=2, value=b.get("slug", ""))
        ws.cell(row=row, column=3, value=b.get("title", ""))
        # column 4 (D) — formula, skip
        ws.cell(row=row, column=5, value=b.get("meta", ""))
        # column 6 (F) — formula, skip
        ws.cell(row=row, column=7, value=b.get("h1", ""))
        # column 8 (H) — formula, skip
        rows_written += 1
    return rows_written


def populate(xlsx_path: str, payload: dict, backup: bool = True) -> dict:
    if backup:
        shutil.copy(xlsx_path, xlsx_path + ".bak")

    wb = load_workbook(xlsx_path, data_only=False)
    _validate_workbook(wb)

    keywords = payload.get("keywords", [])
    pages = payload.get("pages", [])
    briefs = payload.get("briefs", [])

    _check_cannibalization(pages)
    _check_cross_refs(keywords, pages, briefs)

    kw_written = _write_keywords(wb["Keyword Research"], keywords)
    pg_written = _write_pages(wb["Page Map"], pages)
    br_written = _write_briefs(wb["SEO Briefs"], briefs)

    wb.save(xlsx_path)
    return {
        "keywords_written": kw_written,
        "pages_written": pg_written,
        "briefs_written": br_written,
        "backup_path": xlsx_path + ".bak" if backup else None,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Populate SEO_Master.xlsx from a JSON payload.")
    parser.add_argument("--xlsx", required=True, help="Path to SEO_Master.xlsx.")
    parser.add_argument("--payload", required=True, help="Path to payload JSON.")
    parser.add_argument("--no-backup", action="store_true", help="Skip the .bak backup before writing.")
    args = parser.parse_args()

    try:
        with open(args.payload, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except (OSError, json.JSONDecodeError) as exc:
        print(f"Failed to read payload: {exc}", file=sys.stderr)
        return 2

    try:
        report = populate(args.xlsx, payload, backup=not args.no_backup)
    except Exception as exc:
        print(f"populate failed: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1

    print(
        f"Wrote {report['keywords_written']} keywords, "
        f"{report['pages_written']} pages, "
        f"{report['briefs_written']} briefs to {args.xlsx}."
    )
    if report["backup_path"]:
        print(f"Backup saved at {report['backup_path']}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
